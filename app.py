# -*- coding: utf-8 -*-
"""
📊 Personal Finance Dashboard — Streamlit Web App
Connects to "Personal_Finance_Dashboard - Year - Month.xlsx" (10-sheet workbook)
and gives you a live, editable, Excel-linked dashboard.

Run with:  streamlit run app.py
"""

import io
import os
from datetime import datetime, date

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook

# ============================================================================
# 0. PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="Personal Finance Dashboard",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================================
# 1. CUSTOM CSS — Modern Dark-Teal Glassmorphic UI
# ============================================================================
CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"]  {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* ---- App background ---- */
.stApp {
    background: radial-gradient(circle at 10% 0%, #16213a 0%, #0f172a 45%, #0b1220 100%);
    color: #e2f4f1;
}
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #10192e 0%, #0c1424 100%);
    border-right: 1px solid rgba(2, 195, 154, 0.15);
}
section[data-testid="stSidebar"] * { color: #d6f5ef; }

/* ---- Headings ---- */
h1, h2, h3, h4 { color: #eafffb !important; letter-spacing: .2px; }
h1 { font-weight: 800 !important; }
h2, h3 { font-weight: 700 !important; }

/* ---- Tabs ---- */
.stTabs [data-baseweb="tab-list"] { gap: 6px; }
.stTabs [data-baseweb="tab"] {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(2, 195, 154, 0.18);
    border-radius: 12px 12px 0 0;
    padding: 10px 16px;
    color: #9fd8cd;
    font-weight: 600;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, rgba(2,195,154,0.25), rgba(2,128,144,0.25)) !important;
    color: #ffffff !important;
    border-bottom: 2px solid #02c39a !important;
}

/* ---- Metric Cards (custom HTML) ---- */
.metric-card {
    background: linear-gradient(145deg, rgba(30,41,59,0.85), rgba(15,23,42,0.85));
    border: 1px solid rgba(2, 195, 154, 0.28);
    border-radius: 18px;
    padding: 18px 20px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.35), 0 0 18px rgba(2,195,154,0.08) inset;
    backdrop-filter: blur(6px);
    margin-bottom: 10px;
    height: 100%;
    transition: transform .15s ease;
}
.metric-card:hover { transform: translateY(-2px); border-color: rgba(2,195,154,0.55); }
.metric-card .mc-label { font-size: 13px; font-weight: 600; color: #8fd4c7; opacity: .9; margin-bottom: 6px;}
.metric-card .mc-value { font-size: 26px; font-weight: 800; color: #ffffff; line-height: 1.15; }
.metric-card .mc-sub { font-size: 12px; color: #7fa9a0; margin-top: 6px; }
.metric-card .mc-pos { color: #02c39a; font-weight: 700; }
.metric-card .mc-neg { color: #ff6b6b; font-weight: 700; }

.section-title {
    display:flex; align-items:center; gap:8px;
    font-size: 20px; font-weight: 800; margin: 18px 0 8px 0; color:#eafffb;
    border-left: 4px solid #02c39a; padding-left: 10px;
}

/* ---- Buttons ---- */
.stButton>button, .stDownloadButton>button {
    background: linear-gradient(135deg, #02c39a, #028090);
    color: white; border: none; border-radius: 12px; font-weight: 700;
    padding: 10px 18px; box-shadow: 0 4px 14px rgba(2,195,154,0.35);
}
.stButton>button:hover, .stDownloadButton>button:hover {
    filter: brightness(1.1); box-shadow: 0 6px 18px rgba(2,195,154,0.5);
}

/* ---- DataFrames / editors ---- */
[data-testid="stDataFrame"], [data-testid="stDataEditor"] {
    border-radius: 14px; overflow: hidden;
    border: 1px solid rgba(2,195,154,0.2);
}

hr { border-color: rgba(2,195,154,0.15); }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ============================================================================
# 2. SHEET / COLUMN LAYOUT CONSTANTS  (must match the workbook exactly)
# ============================================================================
SH_DASH = "01_Dashboard"
SH_IE = "02_Income_Expense"
SH_CF = "03_Cash_Flow"
SH_EF = "04_Emergency_Fund"
SH_DEBT = "05_Debt"
SH_ASSET = "06_Assets"
SH_NW = "07_Net_Worth"
SH_SI = "08_Saving_Investment"
SH_REPORT = "09_Monthly_Report"
SH_SET = "10_Settings"

IE_COLS = ["Date / ວັນທີ", "Type / ປະເພດ", "Category / ໝວດ", "Description / ລາຍລະອຽດ",
           "Account / ບັນຊີ", "Need_Want / ຈຳເປັນ-ຢາກໄດ້", "Amount / ຈຳນວນເງິນ",
           "Month / ເດືອນ", "Year / ປີ", "Note / ໝາຍເຫດ"]
IE_HEADER_ROW = 2

DEBT_COLS = ["Debt ID / ລະຫັດໜີ້", "Debt Name / ຊື່ໜີ້ສິນ", "Lender / ຜູ້ໃຫ້ກູ້",
             "Original Principal / ຕົ້ນທຶນເດີມ", "Principal Paid / ຊຳລະແລ້ວ",
             "Remaining Principal / ຄົງເຫຼືອ", "Interest Rate / ອັດຕາດອກເບ້ຍ",
             "Monthly Payment / ຊຳລະລາຍເດືອນ", "Due Date / ວັນຄົບກຳນົດ", "Start Date / ວັນເລີ່ມ",
             "End Date / ວັນສິ້ນສຸດ", "Status / ສະຖານະ", "Notes / ໝາຍເຫດ"]
DEBT_HEADER_ROW = 2

ASSET_COLS = ["Asset ID / ລະຫັດ", "Asset Name / ຊື່ຊັບສິນ", "Category / ໝວດ", "Purchase Date / ວັນຊື້",
              "Purchase Cost / ລາຄາຊື້", "Current Value / ມູນຄ່າປັດຈຸບັນ", "Account / ບັນຊີ",
              "Liquidity / ສະພາບຄ່ອງ", "Notes / ໝາຍເຫດ"]
ASSET_HEADER_ROW = 2

EF_COLS = ["Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ", "Target Months / ເປົ້າໝາຍ (ເດືອນ)",
           "Target Fund / ຍອດເປົ້າໝາຍ", "Current Fund / ຍອດປັດຈຸບັນ", "Remaining / ຄົງເຫຼືອ",
           "Progress % / ຄວາມຄືບໜ້າ", "Status / ສະຖານະ"]
EF_HEADER_ROW = 2

SI_LOG_COLS = ["Date / ວັນທີ", "Type / ປະເພດ", "Category / ໝວດ", "Account/Platform / ບັນຊີ-ແພລດຟອມ",
               "Amount / ຈຳນວນເງິນ", "Goal / ເປົ້າໝາຍ", "Notes / ໝາຍເຫດ"]
SI_LOG_HEADER_ROW = 13

SET_COLS = ["Years / ປີ", "Month # / ເລກເດືອນ", "Month Name / ຊື່ເດືອນ", "Income Categories / ໝວດລາຍຮັບ",
            "Expense Categories / ໝວດລາຍຈ່າຍ", "Accounts / ບັນຊີ", "Need_Want / ຈຳເປັນ/ຢາກໄດ້",
            "Debt Status / ສະຖານະໜີ້ສິນ", "Asset Categories / ໝວດຊັບສິນ",
            "Investment Categories / ໝວດການລົງທຶນ", "EF Target Months / ເປົ້າໝາຍ ດ.ສຸກເສີນ",
            "Type / ປະເພດ", "Liquidity / ສະພາບຄ່ອງ", "Saving Category / ໝວດການອອມ",
            "Saving/Investment Type / ປະເພດອອມ/ລົງທຶນ", "_spacer_",
            "All Categories (Income+Expense) — used by Category dropdown"]
SET_HEADER_ROW = 2

LAO_MONTHS = ["ມັງກອນ", "ກຸມພາ", "ມີນາ", "ເມສາ", "ພຶດສະພາ", "ມິຖຸນາ", "ກໍລະກົດ", "ສິງຫາ",
              "ກັນຍາ", "ຕຸລາ", "ພະຈິກ", "ທັນວາ"]

TEAL_PALETTE = ["#02c39a", "#00a896", "#028090", "#05668d", "#5ee6c2", "#7bdff2",
                "#f2b134", "#ff6b6b", "#c9f2e0", "#00e0b8"]

DEFAULT_FILE = "Personal_Finance_Dashboard_-_Year_-_Month.xlsx"


# ============================================================================
# 3. LOAD WORKBOOK -> DATAFRAMES
# ============================================================================
def _read_table(wb, sheet, header_row, columns):
    """Read a rectangular table that starts at `header_row` (1-indexed) using
    fixed column names (so we don't depend on merged/duplicate header cells)."""
    ws = wb[sheet]
    ncols = len(columns)
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c + 1).value for c in range(ncols)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            continue
        rows.append(vals)
    df = pd.DataFrame(rows, columns=columns)
    return df


def load_workbook_data(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    data = {
        "income_expense": _read_table(wb, SH_IE, IE_HEADER_ROW, IE_COLS),
        "debt": _read_table(wb, SH_DEBT, DEBT_HEADER_ROW, DEBT_COLS),
        "assets": _read_table(wb, SH_ASSET, ASSET_HEADER_ROW, ASSET_COLS),
        "ef": _read_table(wb, SH_EF, EF_HEADER_ROW, EF_COLS),
        "saving_log": _read_table(wb, SH_SI, SI_LOG_HEADER_ROW, SI_LOG_COLS),
        "settings": _read_table(wb, SH_SET, SET_HEADER_ROW, SET_COLS),
    }
    # Starting cash balance for month 1 (03_Cash_Flow!B5) — plain input, not a formula
    try:
        start_cash = wb[SH_CF]["B5"].value or 0
    except Exception:
        start_cash = 0
    if not isinstance(start_cash, (int, float)):
        start_cash = 0
    data["start_cash"] = float(start_cash)

    # numeric coercion
    ie = data["income_expense"]
    ie["Amount / ຈຳນວນເງິນ"] = pd.to_numeric(ie["Amount / ຈຳນວນເງິນ"], errors="coerce").fillna(0.0)
    ie["Month / ເດືອນ"] = pd.to_numeric(ie["Month / ເດືອນ"], errors="coerce")
    ie["Year / ປີ"] = pd.to_numeric(ie["Year / ປີ"], errors="coerce")

    debt = data["debt"]
    for c in ["Original Principal / ຕົ້ນທຶນເດີມ", "Principal Paid / ຊຳລະແລ້ວ", "Interest Rate / ອັດຕາດອກເບ້ຍ",
              "Monthly Payment / ຊຳລະລາຍເດືອນ"]:
        debt[c] = pd.to_numeric(debt[c], errors="coerce").fillna(0.0)

    assets = data["assets"]
    for c in ["Purchase Cost / ລາຄາຊື້", "Current Value / ມູນຄ່າປັດຈຸບັນ"]:
        assets[c] = pd.to_numeric(assets[c], errors="coerce").fillna(0.0)

    ef = data["ef"]
    if len(ef) == 0:
        ef = pd.DataFrame([{"Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ": 0.0,
                             "Target Months / ເປົ້າໝາຍ (ເດືອນ)": 3.0}])
        data["ef"] = ef
    ef["Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ"] = pd.to_numeric(
        ef["Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ"], errors="coerce").fillna(0.0)
    ef["Target Months / ເປົ້າໝາຍ (ເດືອນ)"] = pd.to_numeric(
        ef["Target Months / ເປົ້າໝາຍ (ເດືອນ)"], errors="coerce").fillna(0.0)

    silog = data["saving_log"]
    silog["Amount / ຈຳນວນເງິນ"] = pd.to_numeric(silog["Amount / ຈຳນວນເງິນ"], errors="coerce").fillna(0.0)

    return data


# ============================================================================
# 4. FORMULA-EQUIVALENT COMPUTATIONS (mirrors the workbook's named formulas)
# ============================================================================
def sumifs(df, amount_col, **filters):
    d = df
    for col, val in filters.items():
        if val is None:
            continue
        d = d[d[col] == val]
    return float(d[amount_col].sum()) if len(d) else 0.0


def compute_dashboard(dfs, year, month):
    ie = dfs["income_expense"]
    A, TY, TC, TNW, TM, TYR = ("Amount / ຈຳນວນເງິນ", "Type / ປະເພດ", "Category / ໝວດ",
                               "Need_Want / ຈຳເປັນ-ຢາກໄດ້", "Month / ເດືອນ", "Year / ປີ")

    total_income = sumifs(ie, A, **{TY: "Income", TM: month, TYR: year})
    total_expense = sumifs(ie, A, **{TY: "Expense", TM: month, TYR: year})
    cash_flow = total_income - total_expense

    debt = dfs["debt"]
    debt_total_remaining = float((debt["Original Principal / ຕົ້ນທຶນເດີມ"] - debt["Principal Paid / ຊຳລະແລ້ວ"]).sum())
    debt_monthly_payment = float(debt["Monthly Payment / ຊຳລະລາຍເດືອນ"].sum())
    orig_sum = float(debt["Original Principal / ຕົ້ນທຶນເດີມ"].sum())
    paid_sum = float(debt["Principal Paid / ຊຳລະແລ້ວ"].sum())
    debt_paid_pct = (paid_sum / orig_sum) if orig_sum else 0.0

    assets = dfs["assets"]
    asset_total = float(assets["Current Value / ມູນຄ່າປັດຈຸບັນ"].sum())
    asset_ef = float(assets.loc[assets["Category / ໝວດ"] == "Emergency Fund",
                                 "Current Value / ມູນຄ່າປັດຈຸບັນ"].sum())

    net_worth = asset_total - debt_total_remaining

    ef_row = dfs["ef"].iloc[0]
    ef_essential = float(ef_row["Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ"])
    ef_months = float(ef_row["Target Months / ເປົ້າໝາຍ (ເດືອນ)"])
    ef_target = ef_essential * ef_months
    ef_current = asset_ef
    ef_remaining = max(ef_target - ef_current, 0.0)
    ef_progress = (ef_current / ef_target) if ef_target else 0.0
    if ef_target == 0:
        ef_status = "—"
    elif ef_progress >= 1:
        ef_status = "🟢 COMPLETE / ສຳເລັດ"
    elif ef_progress >= 0.75:
        ef_status = "🟢 GOOD / ດີ"
    elif ef_progress >= 0.5:
        ef_status = "🟡 BUILDING / ກຳລັງສະສົມ"
    else:
        ef_status = "🔴 NEED TO BUILD / ຕ້ອງເລັ່ງສະສົມ"

    saving_month = sumifs(ie, A, **{TY: "Expense", TNW: "Saving", TM: month, TYR: year})
    invest_month = sumifs(ie, A, **{TY: "Expense", TNW: "Investment", TM: month, TYR: year})
    saving_all = sumifs(ie, A, **{TY: "Expense", TNW: "Saving"})
    invest_all = sumifs(ie, A, **{TY: "Expense", TNW: "Investment"})
    saving_rate = (saving_month / total_income) if total_income else 0.0
    invest_rate = (invest_month / total_income) if total_income else 0.0

    exp_month = ie[(ie[TY] == "Expense") & (ie[TM] == month) & (ie[TYR] == year)]
    expense_by_cat = exp_month.groupby(TC)[A].sum().sort_values(ascending=False)
    need_want = exp_month.groupby(TNW)[A].sum()

    return dict(
        total_income=total_income, total_expense=total_expense, cash_flow=cash_flow,
        debt_total_remaining=debt_total_remaining, debt_monthly_payment=debt_monthly_payment,
        debt_paid_pct=debt_paid_pct, asset_total=asset_total, asset_ef=asset_ef,
        net_worth=net_worth, ef_essential=ef_essential, ef_months=ef_months, ef_target=ef_target,
        ef_current=ef_current, ef_remaining=ef_remaining, ef_progress=ef_progress, ef_status=ef_status,
        saving_month=saving_month, invest_month=invest_month, saving_all=saving_all, invest_all=invest_all,
        saving_rate=saving_rate, invest_rate=invest_rate,
        expense_by_cat=expense_by_cat, need_want=need_want,
    )


def compute_cash_flow_table(dfs, year, start_cash):
    ie = dfs["income_expense"]
    A, TY, TNW, TM, TYR = ("Amount / ຈຳນວນເງິນ", "Type / ປະເພດ", "Need_Want / ຈຳເປັນ-ຢາກໄດ້",
                            "Month / ເດືອນ", "Year / ປີ")
    rows = []
    beginning = start_cash
    for m in range(1, 13):
        income = sumifs(ie, A, **{TY: "Income", TM: m, TYR: year})
        expense = sumifs(ie, A, **{TY: "Expense", TM: m, TYR: year})
        net = income - expense
        saving = sumifs(ie, A, **{TY: "Expense", TNW: "Saving", TM: m, TYR: year})
        invest = sumifs(ie, A, **{TY: "Expense", TNW: "Investment", TM: m, TYR: year})
        ending = beginning + net
        rows.append({
            "Month": LAO_MONTHS[m - 1], "Beginning Cash": beginning, "Income": income,
            "Expense": expense, "Net Cash Flow": net, "Saving": saving, "Investment": invest,
            "Ending Cash": ending,
        })
        beginning = ending
    return pd.DataFrame(rows)


def compute_monthly_report(dfs, year):
    ie = dfs["income_expense"]
    A, TY, TC, TNW, TM, TYR = ("Amount / ຈຳນວນເງິນ", "Type / ປະເພດ", "Category / ໝວດ",
                               "Need_Want / ຈຳເປັນ-ຢາກໄດ້", "Month / ເດືອນ", "Year / ປີ")
    rows = []
    for m in range(1, 13):
        income = sumifs(ie, A, **{TY: "Income", TM: m, TYR: year})
        expense = sumifs(ie, A, **{TY: "Expense", TM: m, TYR: year})
        cashflow = income - expense
        saving = sumifs(ie, A, **{TY: "Expense", TNW: "Saving", TM: m, TYR: year})
        invest = sumifs(ie, A, **{TY: "Expense", TNW: "Investment", TM: m, TYR: year})
        debt_pay = sumifs(ie, A, **{TY: "Expense", TC: "Debt Payment", TM: m, TYR: year})
        saving_rate = (saving / income) if income else 0.0
        invest_rate = (invest / income) if income else 0.0
        rows.append({
            "Month": LAO_MONTHS[m - 1], "Income": income, "Expense": expense, "Cash Flow": cashflow,
            "Saving": saving, "Investment": invest, "Debt Payment": debt_pay,
            "Saving Rate": saving_rate, "Investment Rate": invest_rate,
        })
    return pd.DataFrame(rows)


# ============================================================================
# 5. EXPORT — write edited data back into the ORIGINAL workbook (keeps formulas)
# ============================================================================
def _set(ws, row, col, value):
    """openpyxl's ws.cell(row, col, value=None) silently NO-OPS when value is
    None (it only assigns when value is not None) — so it can never be used
    to clear a cell. Always set .value explicitly instead."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    return cell


def _clear_range(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            _set(ws, r, c, None)


def _safe(v):
    if v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def build_export_bytes(base_bytes, dfs, selected_year, selected_month, start_cash):
    wb = load_workbook(io.BytesIO(base_bytes), data_only=False)

    # ---- 02_Income_Expense ----
    ws = wb[SH_IE]
    df = dfs["income_expense"]
    _clear_range(ws, IE_HEADER_ROW + 1, max(ws.max_row, IE_HEADER_ROW + len(df) + 5), 1, len(IE_COLS))
    for i, (_, row) in enumerate(df.iterrows()):
        r = IE_HEADER_ROW + 1 + i
        for j, col in enumerate(IE_COLS):
            val = row[col]
            if col in ("Date / ວັນທີ",) and isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            _set(ws, r, j + 1, _safe(val))
    n2 = max(len(df), 1)
    last2 = IE_HEADER_ROW + n2
    for name, letter in [("TxnDate", "A"), ("TxnType", "B"), ("TxnCat", "C"), ("TxnAcc", "E"),
                          ("TxnNW", "F"), ("TxnAmount", "G"), ("TxnMonth", "H"), ("TxnYear", "I")]:
        if name in wb.defined_names:
            wb.defined_names[name].attr_text = f"'{SH_IE}'!${letter}${IE_HEADER_ROW + 1}:${letter}${last2}"

    # ---- 04_Emergency_Fund (only inputs A3 / B3) ----
    ws = wb[SH_EF]
    ef_row = dfs["ef"].iloc[0]
    _set(ws, 3, 1, _safe(float(ef_row["Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ"])))
    _set(ws, 3, 2, _safe(float(ef_row["Target Months / ເປົ້າໝາຍ (ເດືອນ)"])))

    # ---- 05_Debt ----
    ws = wb[SH_DEBT]
    df = dfs["debt"]
    _clear_range(ws, DEBT_HEADER_ROW + 1, max(ws.max_row, DEBT_HEADER_ROW + len(df) + 5), 1, len(DEBT_COLS))
    for i, (_, row) in enumerate(df.iterrows()):
        r = DEBT_HEADER_ROW + 1 + i
        for j, col in enumerate(DEBT_COLS):
            if col == "Remaining Principal / ຄົງເຫຼືອ":
                _set(ws, r, j + 1, f'=IF(D{r}="","",D{r}-E{r})')
                continue
            if col == "Status / ສະຖານະ":
                _set(ws, r, j + 1,
                     (f'=IF(D{r}="","",IF(F{r}<=0,"✅ PAID / ຈ່າຍໝົດ",'
                      f'IF(AND(I{r}<>"",I{r}<TODAY()),"⚠ OVERDUE / ເກີນກຳນົດ",'
                      f'"🔵 ACTIVE / ດຳເນີນຢູ່")))'))
                continue
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            _set(ws, r, j + 1, _safe(val))
    n5 = max(len(df), 1)
    last5 = DEBT_HEADER_ROW + n5
    ws["B37"] = f"=SUM(F{DEBT_HEADER_ROW+1}:F{last5})"
    ws["B38"] = f"=SUM(H{DEBT_HEADER_ROW+1}:H{last5})"
    ws["B41"] = f"=IFERROR(SUM(E{DEBT_HEADER_ROW+1}:E{last5})/SUM(D{DEBT_HEADER_ROW+1}:D{last5}),0)"

    # ---- 06_Assets ----
    ws = wb[SH_ASSET]
    df = dfs["assets"]
    _clear_range(ws, ASSET_HEADER_ROW + 1, max(ws.max_row, ASSET_HEADER_ROW + len(df) + 5), 1, len(ASSET_COLS))
    for i, (_, row) in enumerate(df.iterrows()):
        r = ASSET_HEADER_ROW + 1 + i
        for j, col in enumerate(ASSET_COLS):
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            _set(ws, r, j + 1, _safe(val))
    n6 = max(len(df), 1)
    last6 = ASSET_HEADER_ROW + n6
    ws["B54"] = f"=SUMIFS($F${ASSET_HEADER_ROW+1}:$F${last6},$C${ASSET_HEADER_ROW+1}:$C${last6},A54)"
    ws["B60"] = f"=SUM(F{ASSET_HEADER_ROW+1}:F{last6})"

    # ---- 08_Saving_Investment allocation log ----
    ws = wb[SH_SI]
    df = dfs["saving_log"]
    _clear_range(ws, SI_LOG_HEADER_ROW + 1, max(ws.max_row, SI_LOG_HEADER_ROW + len(df) + 5), 1, len(SI_LOG_COLS))
    for i, (_, row) in enumerate(df.iterrows()):
        r = SI_LOG_HEADER_ROW + 1 + i
        for j, col in enumerate(SI_LOG_COLS):
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            _set(ws, r, j + 1, _safe(val))

    # ---- 10_Settings ----
    ws = wb[SH_SET]
    df = dfs["settings"]
    real_cols = [c for c in SET_COLS if c != "_spacer_"]
    _clear_range(ws, SET_HEADER_ROW + 1, max(ws.max_row, SET_HEADER_ROW + len(df) + 5), 1, len(SET_COLS))
    for i, (_, row) in enumerate(df.iterrows()):
        r = SET_HEADER_ROW + 1 + i
        for j, col in enumerate(SET_COLS):
            if col == "_spacer_":
                continue
            _set(ws, r, j + 1, _safe(row[col]))

    # ---- Selectors & starting cash ----
    wb[SH_DASH]["D3"] = int(selected_year)
    wb[SH_DASH]["D4"] = int(selected_month)
    wb[SH_CF]["B5"] = _safe(float(start_cash))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================================
# 6. UI HELPERS
# ============================================================================
def fmt_money(x):
    try:
        return f"{x:,.0f} ₭"
    except Exception:
        return str(x)


def fmt_pct(x):
    try:
        return f"{x*100:,.1f}%"
    except Exception:
        return str(x)


def metric_card(label, value, sub=None, icon="", positive=None):
    sub_html = ""
    if sub:
        cls = ""
        if positive is True:
            cls = "mc-pos"
        elif positive is False:
            cls = "mc-neg"
        sub_html = f'<div class="mc-sub {cls}">{sub}</div>'
    st.markdown(
        f"""<div class="metric-card">
                <div class="mc-label">{icon} {label}</div>
                <div class="mc-value">{value}</div>
                {sub_html}
            </div>""",
        unsafe_allow_html=True,
    )


def section_title(text):
    st.markdown(f'<div class="section-title">{text}</div>', unsafe_allow_html=True)


def plotly_theme(fig, title=None, height=360):
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", color="#e2f4f1"),
        title=title,
        height=height,
        margin=dict(l=10, r=10, t=50 if title else 20, b=10),
        legend=dict(bgcolor="rgba(0,0,0,0)"),
    )
    return fig


# ============================================================================
# 7. SESSION STATE / DATA LOADING
# ============================================================================
def init_state_from_bytes(file_bytes, filename):
    dfs = load_workbook_data(file_bytes)
    st.session_state.base_bytes = file_bytes
    st.session_state.file_name = filename
    st.session_state.df_ie = dfs["income_expense"]
    st.session_state.df_debt = dfs["debt"]
    st.session_state.df_assets = dfs["assets"]
    st.session_state.df_ef = dfs["ef"]
    st.session_state.df_silog = dfs["saving_log"]
    st.session_state.df_settings = dfs["settings"]
    st.session_state.start_cash = dfs["start_cash"]
    st.session_state.loaded = True


st.sidebar.markdown("## 💰 Finance Dashboard")
st.sidebar.caption("Excel-linked · Modern Dark-Teal UI")

uploaded = st.sidebar.file_uploader("📂 Upload Excel File (.xlsx)", type=["xlsx"])

if uploaded is not None and st.session_state.get("file_name") != uploaded.name:
    init_state_from_bytes(uploaded.read(), uploaded.name)
elif not st.session_state.get("loaded"):
    if os.path.exists(DEFAULT_FILE):
        with open(DEFAULT_FILE, "rb") as f:
            init_state_from_bytes(f.read(), DEFAULT_FILE)
    else:
        st.info("👋 Upload your **Personal_Finance_Dashboard - Year - Month.xlsx** file "
                "from the sidebar to get started.")
        st.stop()

dfs = {
    "income_expense": st.session_state.df_ie,
    "debt": st.session_state.df_debt,
    "assets": st.session_state.df_assets,
    "ef": st.session_state.df_ef,
    "saving_log": st.session_state.df_silog,
    "settings": st.session_state.df_settings,
}

# ---- dropdown option lists (from Settings sheet) ----
set_df = st.session_state.df_settings


def _opts(col):
    if col not in set_df.columns:
        return []
    return sorted([v for v in set_df[col].dropna().unique().tolist() if str(v).strip() != ""])


years_list = _opts("Years / ປີ") or [2025.0, 2026.0, 2027.0]
years_list = sorted({int(y) for y in years_list})
income_cats = _opts("Income Categories / ໝວດລາຍຮັບ")
expense_cats = _opts("Expense Categories / ໝວດລາຍຈ່າຍ")
all_cats = _opts("All Categories (Income+Expense) — used by Category dropdown") or sorted(set(income_cats + expense_cats))
accounts_list = _opts("Accounts / ບັນຊີ")
need_want_list = _opts("Need_Want / ຈຳເປັນ/ຢາກໄດ້")
asset_cats = _opts("Asset Categories / ໝວດຊັບສິນ")
liquidity_list = _opts("Liquidity / ສະພາບຄ່ອງ")
debt_status_list = _opts("Debt Status / ສະຖານະໜີ້ສິນ")
saving_type_list = _opts("Saving/Investment Type / ປະເພດອອມ/ລົງທຶນ") or ["Saving", "Investment"]
saving_cat_list = _opts("Saving Category / ໝວດການອອມ")

# ---- sidebar filters ----
st.sidebar.markdown("---")
st.sidebar.markdown("### 🗓️ Period Filter")
today = date.today()
default_year = today.year if today.year in years_list else (years_list[-1] if years_list else today.year)
sel_year = st.sidebar.selectbox("ປີ (Year)", years_list, index=years_list.index(default_year) if default_year in years_list else 0)
sel_month = st.sidebar.selectbox("ເດືອນ (Month)", list(range(1, 13)),
                                  index=min(today.month, 12) - 1,
                                  format_func=lambda m: f"{m:02d} — {LAO_MONTHS[m-1]}")

st.sidebar.markdown("---")
st.sidebar.markdown("### 💵 Starting Cash Balance")
st.session_state.start_cash = st.sidebar.number_input(
    "Beginning cash (Month 1) / ເງິນສົດຕົ້ນປີ", min_value=0.0,
    value=float(st.session_state.start_cash), step=100000.0, format="%.0f",
)

st.sidebar.markdown("---")


# ============================================================================
# 8. TITLE
# ============================================================================
st.markdown("# 📊 Personal Finance Dashboard")
st.caption(f"ຕິດຕາມການເງິນສ່ວນຕົວ · Income · Expense · Cash Flow · Net Worth &nbsp;|&nbsp; "
           f"Currency: **LAK (₭)** &nbsp;|&nbsp; Period: **{LAO_MONTHS[sel_month-1]} / {sel_year}**")

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 Dashboard Overview",
    "💵 Income & Expense",
    "💳 Debt & Assets",
    "🛟 Emergency & Savings",
    "📅 Monthly Report",
    "⚙️ Settings",
])

# ============================================================================
# TAB 1 — DASHBOARD OVERVIEW
# ============================================================================
with tab1:
    m = compute_dashboard(dfs, sel_year, sel_month)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        metric_card("Total Income / ລາຍຮັບລວມ", fmt_money(m["total_income"]), icon="💰")
    with c2:
        metric_card("Total Expense / ລາຍຈ່າຍລວມ", fmt_money(m["total_expense"]), icon="💸")
    with c3:
        cf = m["cash_flow"]
        metric_card("Cash Flow / ກະແສເງິນສົດ", fmt_money(cf), icon="🔄",
                     sub=("⚠ Overspending / ລາຍຈ່າຍເກີນລາຍຮັບ" if cf < 0 else "✅ On track"),
                     positive=(cf >= 0))
    with c4:
        metric_card("Emergency Fund / ເງິນສຳຮອງສຸກເສີນ", fmt_money(m["ef_current"]), icon="🛟",
                     sub=f"Target: {fmt_money(m['ef_target'])} · {fmt_pct(m['ef_progress'])}")

    c5, c6, c7, c8 = st.columns(4)
    with c5:
        metric_card("Total Debt / ໜີ້ສິນລວມ", fmt_money(m["debt_total_remaining"]), icon="💳",
                     sub=f"Monthly: {fmt_money(m['debt_monthly_payment'])} · Paid {fmt_pct(m['debt_paid_pct'])}")
    with c6:
        metric_card("Total Assets / ຊັບສິນລວມ", fmt_money(m["asset_total"]), icon="🏦",
                     sub=f"Emergency Fund: {fmt_money(m['asset_ef'])}")
    with c7:
        metric_card("Net Worth / ມູນຄ່າສຸດທິ", fmt_money(m["net_worth"]), icon="📊")
    with c8:
        metric_card("Saving & Investment / ອອມ ແລະ ລົງທຶນ",
                     fmt_money(m["saving_month"] + m["invest_month"]), icon="💰",
                     sub=f"Saving {fmt_pct(m['saving_rate'])} · Invest {fmt_pct(m['invest_rate'])}")

    section_title("📈 Charts / ຕາຕະລາງກາຟຟິກ")
    cc1, cc2 = st.columns(2)
    with cc1:
        if len(m["expense_by_cat"]) > 0:
            fig = go.Figure(go.Pie(
                labels=m["expense_by_cat"].index, values=m["expense_by_cat"].values, hole=0.55,
                marker=dict(colors=TEAL_PALETTE, line=dict(color="#0f172a", width=2)),
                textinfo="label+percent",
            ))
            plotly_theme(fig, title="Expense by Category / ລາຍຈ່າຍຕາມໝວດ")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No expense data for this period yet.")
    with cc2:
        if len(m["need_want"]) > 0:
            fig = go.Figure(go.Pie(
                labels=m["need_want"].index, values=m["need_want"].values, hole=0.55,
                marker=dict(colors=["#02c39a", "#f2b134", "#05668d", "#ff6b6b"], line=dict(color="#0f172a", width=2)),
                textinfo="label+percent",
            ))
            plotly_theme(fig, title="Need vs Want / ຈຳເປັນ vs ຢາກໄດ້")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No Need/Want tagged expenses for this period yet.")

    cc3, cc4 = st.columns(2)
    with cc3:
        gauge = go.Figure(go.Indicator(
            mode="gauge+number", value=round(m["ef_progress"] * 100, 1),
            number={"suffix": "%"},
            title={"text": "Emergency Fund Progress / ຄວາມຄືບໜ້າ"},
            gauge={
                "axis": {"range": [0, 100], "tickcolor": "#8fd4c7"},
                "bar": {"color": "#02c39a"},
                "bgcolor": "rgba(255,255,255,0.05)",
                "steps": [
                    {"range": [0, 50], "color": "rgba(255,107,107,0.25)"},
                    {"range": [50, 75], "color": "rgba(242,177,52,0.25)"},
                    {"range": [75, 100], "color": "rgba(2,195,154,0.25)"},
                ],
                "threshold": {"line": {"color": "white", "width": 3}, "value": 100},
            },
        ))
        plotly_theme(gauge, height=320)
        st.plotly_chart(gauge, use_container_width=True)
    with cc4:
        funnel = go.Figure(go.Funnel(
            y=["Income", "Expense", "Net Cash Flow", "Saving + Investment"],
            x=[max(m["total_income"], 0), max(m["total_expense"], 0),
               max(m["cash_flow"], 0), max(m["saving_month"] + m["invest_month"], 0)],
            marker={"color": ["#02c39a", "#028090", "#00a896", "#5ee6c2"]},
        ))
        plotly_theme(funnel, title="Income → Expense → Cash Flow → Saving/Invest", height=320)
        st.plotly_chart(funnel, use_container_width=True)

    section_title("🔄 Cash Flow Trend / ແນວໂນ້ມກະແສເງິນສົດ")
    cf_table = compute_cash_flow_table(dfs, sel_year, st.session_state.start_cash)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=cf_table["Month"], y=cf_table["Income"], name="Income",
                              mode="lines", line=dict(color="#02c39a", width=3), fill="tozeroy",
                              fillcolor="rgba(2,195,154,0.15)"))
    fig.add_trace(go.Scatter(x=cf_table["Month"], y=cf_table["Expense"], name="Expense",
                              mode="lines", line=dict(color="#ff6b6b", width=3), fill="tozeroy",
                              fillcolor="rgba(255,107,107,0.12)"))
    fig.add_trace(go.Scatter(x=cf_table["Month"], y=cf_table["Ending Cash"], name="Ending Cash",
                              mode="lines+markers", line=dict(color="#7bdff2", width=2, dash="dot")))
    plotly_theme(fig, height=380)
    st.plotly_chart(fig, use_container_width=True)

    section_title("💳 Debt Progress / ຄວາມຄືບໜ້າຫນີ້ສິນ")
    debt_df = dfs["debt"]
    if len(debt_df):
        paid = debt_df["Principal Paid / ຊຳລະແລ້ວ"]
        remaining = debt_df["Original Principal / ຕົ້ນທຶນເດີມ"] - paid
        fig = go.Figure()
        fig.add_trace(go.Bar(y=debt_df["Debt Name / ຊື່ໜີ້ສິນ"], x=paid, name="Paid",
                              orientation="h", marker_color="#02c39a"))
        fig.add_trace(go.Bar(y=debt_df["Debt Name / ຊື່ໜີ້ສິນ"], x=remaining, name="Remaining",
                              orientation="h", marker_color="#ff6b6b"))
        fig.update_layout(barmode="stack")
        plotly_theme(fig, height=max(300, 60 * len(debt_df)))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No debt records yet.")

# ============================================================================
# TAB 2 — INCOME & EXPENSE EDITOR
# ============================================================================
with tab2:
    section_title("💵 Income & Expense Transaction Log / ບັນທຶກທຸລະກຳ")
    st.caption("Double-click any cell to edit. Use the **+** row at the bottom to add a new transaction, "
               "or select a row and press delete to remove it.")

    edited_ie = st.data_editor(
        st.session_state.df_ie,
        num_rows="dynamic",
        use_container_width=True,
        height=460,
        column_config={
            "Date / ວັນທີ": st.column_config.DateColumn("Date", format="YYYY-MM-DD"),
            "Type / ປະເພດ": st.column_config.SelectboxColumn("Type", options=["Income", "Expense"]),
            "Category / ໝວດ": st.column_config.SelectboxColumn("Category", options=all_cats or None),
            "Description / ລາຍລະອຽດ": st.column_config.TextColumn("Description"),
            "Account / ບັນຊີ": st.column_config.SelectboxColumn("Account", options=accounts_list or None),
            "Need_Want / ຈຳເປັນ-ຢາກໄດ້": st.column_config.SelectboxColumn("Need/Want/Saving/Invest",
                                                                          options=need_want_list or None),
            "Amount / ຈຳນວນເງິນ": st.column_config.NumberColumn("Amount (₭)", format="%.0f"),
            "Month / ເດືອນ": st.column_config.NumberColumn("Month", min_value=1, max_value=12, step=1),
            "Year / ປີ": st.column_config.NumberColumn("Year", step=1),
            "Note / ໝາຍເຫດ": st.column_config.TextColumn("Note"),
        },
        key="editor_ie",
    )
    # auto-fill Month/Year from Date when possible, keep numeric types clean
    edited_ie = edited_ie.copy()
    if "Date / ວັນທີ" in edited_ie.columns:
        dts = pd.to_datetime(edited_ie["Date / ວັນທີ"], errors="coerce")
        edited_ie["Month / ເດືອນ"] = np.where(dts.notna(), dts.dt.month, edited_ie["Month / ເດືອນ"])
        edited_ie["Year / ປີ"] = np.where(dts.notna(), dts.dt.year, edited_ie["Year / ປີ"])
    edited_ie["Amount / ຈຳນວນເງິນ"] = pd.to_numeric(edited_ie["Amount / ຈຳນວນເງິນ"], errors="coerce").fillna(0.0)
    st.session_state.df_ie = edited_ie
    dfs["income_expense"] = edited_ie

    m2 = compute_dashboard(dfs, sel_year, sel_month)
    q1, q2, q3 = st.columns(3)
    with q1:
        metric_card("Income (selected period)", fmt_money(m2["total_income"]), icon="💰")
    with q2:
        metric_card("Expense (selected period)", fmt_money(m2["total_expense"]), icon="💸")
    with q3:
        metric_card("Net Cash Flow", fmt_money(m2["cash_flow"]), icon="🔄", positive=(m2["cash_flow"] >= 0))

    if len(edited_ie):
        by_type = edited_ie[(edited_ie["Year / ປີ"] == sel_year) & (edited_ie["Month / ເດືອນ"] == sel_month)]
        if len(by_type):
            fig = px.bar(by_type.groupby("Type / ປະເພດ")["Amount / ຈຳນວນເງິນ"].sum().reset_index(),
                         x="Type / ປະເພດ", y="Amount / ຈຳນວນເງິນ", color="Type / ປະເພດ",
                         color_discrete_sequence=["#02c39a", "#ff6b6b"])
            plotly_theme(fig, title="Income vs Expense (selected period)", height=320)
            st.plotly_chart(fig, use_container_width=True)

# ============================================================================
# TAB 3 — DEBT & ASSETS TRACKER
# ============================================================================
with tab3:
    section_title("💳 Debt Tracker / ຕິດຕາມໜີ້ສິນ")
    debt_view = st.session_state.df_debt.copy()
    debt_view["Remaining Principal / ຄົງເຫຼືອ"] = (
        debt_view["Original Principal / ຕົ້ນທຶນເດີມ"] - debt_view["Principal Paid / ຊຳລະແລ້ວ"])

    edited_debt = st.data_editor(
        debt_view,
        num_rows="dynamic",
        use_container_width=True,
        height=320,
        column_config={
            "Original Principal / ຕົ້ນທຶນເດີມ": st.column_config.NumberColumn("Original Principal (₭)", format="%.0f"),
            "Principal Paid / ຊຳລະແລ້ວ": st.column_config.NumberColumn("Principal Paid (₭)", format="%.0f"),
            "Remaining Principal / ຄົງເຫຼືອ": st.column_config.NumberColumn("Remaining (auto)", format="%.0f",
                                                                            disabled=True),
            "Interest Rate / ອັດຕາດອກເບ້ຍ": st.column_config.NumberColumn("Interest Rate", format="%.4f"),
            "Monthly Payment / ຊຳລະລາຍເດືອນ": st.column_config.NumberColumn("Monthly Payment (₭)", format="%.0f"),
            "Due Date / ວັນຄົບກຳນົດ": st.column_config.DateColumn("Due Date"),
            "Start Date / ວັນເລີ່ມ": st.column_config.DateColumn("Start Date"),
            "End Date / ວັນສິ້ນສຸດ": st.column_config.DateColumn("End Date"),
            "Status / ສະຖານະ": st.column_config.SelectboxColumn("Status", options=debt_status_list or None),
        },
        key="editor_debt",
    )
    edited_debt = edited_debt.copy()
    edited_debt["Original Principal / ຕົ້ນທຶນເດີມ"] = pd.to_numeric(
        edited_debt["Original Principal / ຕົ້ນທຶນເດີມ"], errors="coerce").fillna(0.0)
    edited_debt["Principal Paid / ຊຳລະແລ້ວ"] = pd.to_numeric(
        edited_debt["Principal Paid / ຊຳລະແລ້ວ"], errors="coerce").fillna(0.0)
    edited_debt["Remaining Principal / ຄົງເຫຼືອ"] = (
        edited_debt["Original Principal / ຕົ້ນທຶນເດີມ"] - edited_debt["Principal Paid / ຊຳລະແລ້ວ"])
    st.session_state.df_debt = edited_debt.drop(columns=["Remaining Principal / ຄົງເຫຼືອ"]).assign(
        **{"Remaining Principal / ຄົງເຫຼືອ": edited_debt["Remaining Principal / ຄົງເຫຼືອ"]}
    )[DEBT_COLS]
    dfs["debt"] = st.session_state.df_debt

    section_title("🏦 Asset Tracker / ຕິດຕາມຊັບສິນ")
    edited_assets = st.data_editor(
        st.session_state.df_assets,
        num_rows="dynamic",
        use_container_width=True,
        height=320,
        column_config={
            "Category / ໝວດ": st.column_config.SelectboxColumn("Category", options=asset_cats or None),
            "Purchase Date / ວັນຊື້": st.column_config.DateColumn("Purchase Date"),
            "Purchase Cost / ລາຄາຊື້": st.column_config.NumberColumn("Purchase Cost (₭)", format="%.0f"),
            "Current Value / ມູນຄ່າປັດຈຸບັນ": st.column_config.NumberColumn("Current Value (₭)", format="%.0f"),
            "Account / ບັນຊີ": st.column_config.SelectboxColumn("Account", options=accounts_list or None),
            "Liquidity / ສະພາບຄ່ອງ": st.column_config.SelectboxColumn("Liquidity", options=liquidity_list or None),
        },
        key="editor_assets",
    )
    edited_assets = edited_assets.copy()
    edited_assets["Current Value / ມູນຄ່າປັດຈຸບັນ"] = pd.to_numeric(
        edited_assets["Current Value / ມູນຄ່າປັດຈຸບັນ"], errors="coerce").fillna(0.0)
    st.session_state.df_assets = edited_assets
    dfs["assets"] = edited_assets

    m3 = compute_dashboard(dfs, sel_year, sel_month)
    r1, r2, r3 = st.columns(3)
    with r1:
        metric_card("Total Debt (Remaining)", fmt_money(m3["debt_total_remaining"]), icon="💳")
    with r2:
        metric_card("Total Assets", fmt_money(m3["asset_total"]), icon="🏦")
    with r3:
        metric_card("Net Worth", fmt_money(m3["net_worth"]), icon="📊",
                     positive=(m3["net_worth"] >= 0))

    cc1, cc2 = st.columns(2)
    with cc1:
        if len(edited_debt):
            fig = go.Figure()
            fig.add_trace(go.Bar(y=edited_debt["Debt Name / ຊື່ໜີ້ສິນ"],
                                  x=edited_debt["Principal Paid / ຊຳລະແລ້ວ"],
                                  name="Paid", orientation="h", marker_color="#02c39a"))
            fig.add_trace(go.Bar(y=edited_debt["Debt Name / ຊື່ໜີ້ສິນ"],
                                  x=edited_debt["Remaining Principal / ຄົງເຫຼືອ"],
                                  name="Remaining", orientation="h", marker_color="#ff6b6b"))
            fig.update_layout(barmode="stack")
            plotly_theme(fig, title="Debt Progress", height=max(300, 55 * len(edited_debt)))
            st.plotly_chart(fig, use_container_width=True)
    with cc2:
        if len(edited_assets):
            by_cat = edited_assets.groupby("Category / ໝວດ")["Current Value / ມູນຄ່າປັດຈຸບັນ"].sum()
            fig = go.Figure(go.Pie(labels=by_cat.index, values=by_cat.values, hole=0.55,
                                    marker=dict(colors=TEAL_PALETTE, line=dict(color="#0f172a", width=2)),
                                    textinfo="label+percent"))
            plotly_theme(fig, title="Assets by Category", height=max(300, 55 * len(edited_assets)))
            st.plotly_chart(fig, use_container_width=True)

# ============================================================================
# TAB 4 — EMERGENCY FUND & SAVINGS
# ============================================================================
with tab4:
    section_title("🛟 Emergency Fund / ເງິນສຳຮອງສຸກເສີນ")
    ef_row = st.session_state.df_ef.iloc[0]
    ec1, ec2 = st.columns(2)
    with ec1:
        new_essential = st.number_input(
            "Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ (₭)",
            min_value=0.0, value=float(ef_row["Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ"]),
            step=100000.0, format="%.0f",
        )
    with ec2:
        new_months = st.number_input(
            "Target Months / ເປົ້າໝາຍ (ເດືອນ)", min_value=0.0,
            value=float(ef_row["Target Months / ເປົ້າໝາຍ (ເດືອນ)"]), step=1.0,
        )
    st.session_state.df_ef = pd.DataFrame([{
        "Monthly Essential Expense / ລາຍຈ່າຍຈຳເປັນຕໍ່ເດືອນ": new_essential,
        "Target Months / ເປົ້າໝາຍ (ເດືອນ)": new_months,
    }])
    dfs["ef"] = st.session_state.df_ef

    m4 = compute_dashboard(dfs, sel_year, sel_month)
    e1, e2, e3, e4 = st.columns(4)
    with e1:
        metric_card("Target Fund", fmt_money(m4["ef_target"]), icon="🎯")
    with e2:
        metric_card("Current Fund", fmt_money(m4["ef_current"]), icon="🛟")
    with e3:
        metric_card("Remaining", fmt_money(m4["ef_remaining"]), icon="📉")
    with e4:
        metric_card("Status", m4["ef_status"], icon="📌")

    gauge = go.Figure(go.Indicator(
        mode="gauge+number", value=round(m4["ef_progress"] * 100, 1), number={"suffix": "%"},
        title={"text": "Progress / ຄວາມຄືບໜ້າ"},
        gauge={"axis": {"range": [0, 100]}, "bar": {"color": "#02c39a"},
               "bgcolor": "rgba(255,255,255,0.05)",
               "steps": [{"range": [0, 50], "color": "rgba(255,107,107,0.25)"},
                         {"range": [50, 75], "color": "rgba(242,177,52,0.25)"},
                         {"range": [75, 100], "color": "rgba(2,195,154,0.25)"}]},
    ))
    plotly_theme(gauge, height=300)
    st.plotly_chart(gauge, use_container_width=True)
    st.caption("💡 Current Fund is pulled automatically from **06_Assets** where "
               "Category = *Emergency Fund*. Edit that amount in the Debt & Assets tab.")

    st.markdown("---")
    section_title("💰 Saving & Investment / ເງິນອອມ ແລະ ການລົງທຶນ")

    s1, s2, s3, s4 = st.columns(4)
    with s1:
        metric_card("Total Saving (all-time)", fmt_money(m4["saving_all"]), icon="💰")
    with s2:
        metric_card("Total Investment (all-time)", fmt_money(m4["invest_all"]), icon="📈")
    with s3:
        metric_card("Monthly Saving Rate", fmt_pct(m4["saving_rate"]), icon="🧮")
    with s4:
        metric_card("Monthly Investment Rate", fmt_pct(m4["invest_rate"]), icon="🧮")

    st.markdown("##### 📒 Allocation Log / ບັນທຶກການຈັດສັນເງິນ")
    st.caption("Tracks *where* saving/investment cash was directed — not a second cash-flow entry.")
    edited_silog = st.data_editor(
        st.session_state.df_silog,
        num_rows="dynamic",
        use_container_width=True,
        height=280,
        column_config={
            "Date / ວັນທີ": st.column_config.DateColumn("Date"),
            "Type / ປະເພດ": st.column_config.SelectboxColumn("Type", options=saving_type_list or None),
            "Category / ໝວດ": st.column_config.SelectboxColumn("Category", options=saving_cat_list or None),
            "Amount / ຈຳນວນເງິນ": st.column_config.NumberColumn("Amount (₭)", format="%.0f"),
        },
        key="editor_silog",
    )
    edited_silog = edited_silog.copy()
    edited_silog["Amount / ຈຳນວນເງິນ"] = pd.to_numeric(edited_silog["Amount / ຈຳນວນເງິນ"], errors="coerce").fillna(0.0)
    st.session_state.df_silog = edited_silog
    dfs["saving_log"] = edited_silog

# ============================================================================
# TAB 5 — MONTHLY FINANCIAL REPORT
# ============================================================================
with tab5:
    section_title(f"📅 Monthly Financial Report — {sel_year}")
    report = compute_monthly_report(dfs, sel_year)

    display = report.copy()
    for c in ["Income", "Expense", "Cash Flow", "Saving", "Investment", "Debt Payment"]:
        display[c] = display[c].map(fmt_money)
    for c in ["Saving Rate", "Investment Rate"]:
        display[c] = report[c].map(fmt_pct)
    st.dataframe(display, use_container_width=True, height=460, hide_index=True)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=report["Month"], y=report["Income"], name="Income",
                              mode="lines+markers", line=dict(color="#02c39a", width=3)))
    fig.add_trace(go.Scatter(x=report["Month"], y=report["Expense"], name="Expense",
                              mode="lines+markers", line=dict(color="#ff6b6b", width=3)))
    fig.add_trace(go.Bar(x=report["Month"], y=report["Cash Flow"], name="Cash Flow",
                          marker_color="rgba(2,195,154,0.35)"))
    plotly_theme(fig, title="Income / Expense / Cash Flow Trend", height=380)
    st.plotly_chart(fig, use_container_width=True)

    fig2 = go.Figure()
    fig2.add_trace(go.Bar(x=report["Month"], y=report["Saving"], name="Saving", marker_color="#02c39a"))
    fig2.add_trace(go.Bar(x=report["Month"], y=report["Investment"], name="Investment", marker_color="#028090"))
    fig2.add_trace(go.Bar(x=report["Month"], y=report["Debt Payment"], name="Debt Payment", marker_color="#ff6b6b"))
    fig2.update_layout(barmode="group")
    plotly_theme(fig2, title="Saving vs Investment vs Debt Payment", height=380)
    st.plotly_chart(fig2, use_container_width=True)

    cur = compute_dashboard(dfs, sel_year, sel_month)
    st.markdown("##### 📊 Snapshot at selected period")
    n1, n2, n3 = st.columns(3)
    with n1:
        metric_card("Total Debt (current)", fmt_money(cur["debt_total_remaining"]), icon="💳")
    with n2:
        metric_card("Total Assets (current)", fmt_money(cur["asset_total"]), icon="🏦")
    with n3:
        metric_card("Net Worth (current)", fmt_money(cur["net_worth"]), icon="📊")

# ============================================================================
# TAB 6 — SYSTEM SETTINGS
# ============================================================================
with tab6:
    section_title("⚙️ Settings — Dropdown Lists & Lookup Tables")
    st.caption("These lists power the dropdown menus used throughout the app "
               "(categories, accounts, statuses, etc.). Do not reorder Years / Month columns.")
    settings_view = st.session_state.df_settings.drop(columns=["_spacer_"], errors="ignore")
    edited_settings = st.data_editor(
        settings_view,
        num_rows="dynamic",
        use_container_width=True,
        height=520,
        key="editor_settings",
    )
    if "_spacer_" not in edited_settings.columns:
        edited_settings["_spacer_"] = None
    st.session_state.df_settings = edited_settings[SET_COLS]
    dfs["settings"] = st.session_state.df_settings

# ============================================================================
# 9. SAVE / EXPORT (sidebar)
# ============================================================================
st.sidebar.markdown("### 💾 Save & Export")
try:
    export_bytes = build_export_bytes(
        st.session_state.base_bytes,
        {
            "income_expense": st.session_state.df_ie,
            "debt": st.session_state.df_debt,
            "assets": st.session_state.df_assets,
            "ef": st.session_state.df_ef,
            "saving_log": st.session_state.df_silog,
            "settings": st.session_state.df_settings,
        },
        sel_year, sel_month, st.session_state.start_cash,
    )
    st.sidebar.download_button(
        "💾 ບັນທຶກ/ດາວໂຫລດ Excel",
        data=export_bytes,
        file_name=f"Personal_Finance_Dashboard_{sel_year}_{sel_month:02d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.sidebar.caption("✅ All formulas (KPIs, Dashboard, Net Worth, Monthly Report) "
                       "recalculate automatically when the file is opened in Excel.")
except Exception as e:
    st.sidebar.error(f"Export failed: {e}")

st.sidebar.markdown("---")
st.sidebar.caption("Built with Streamlit · Pandas · Plotly · Openpyxl")
